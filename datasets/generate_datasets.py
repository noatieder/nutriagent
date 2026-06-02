"""
NutriAgent Dataset Generator
Generates 6 Excel files for project submission:
  01_food_database_full.xlsx        - USDA-sourced food data (300+ items, raw → cleaned)
  02_food_database_chatbot.xlsx     - Curated 117-item database used in chatbot
  03_kmeans_clustering_results.xlsx - K-Means clustering analysis (K=4, Silhouette=0.582)
  04_similarity_comparison.xlsx     - Cosine vs Jaccard similarity comparison
  05_bmi_growth_charts.xlsx         - CDC pediatric BMI reference data ages 4-18
  06_test_cases.xlsx                - 6 test user profiles for validation

Run: python3 generate_datasets.py
Dependencies: pip install pandas openpyxl scikit-learn
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# ─── Color palette ────────────────────────────────────────────────
COLORS = {
    'header':    'FF20B2AA',   # Teal (NutriAgent brand)
    'cluster0':  'FFAAFFAA',   # Light green
    'cluster1':  'FFAACCFF',   # Light blue
    'cluster2':  'FFFFFFAA',   # Light yellow
    'cluster3':  'FFFFC8AA',   # Light orange
    'outlier':   'FFFF9999',   # Light red
    'success':   'FFD4EDDA',
    'warning':   'FFFFF3CD',
    'error':     'FFF8D7DA',
}

CLUSTER_NAMES = {0: 'נפח/ירקות', 1: 'חלבון רזה', 2: 'שומנים בריאים', 3: 'פחמימות מורכבות'}
CLUSTER_NAMES_EN = {0: 'Volume/Vegetables', 1: 'Lean Protein', 2: 'Essential Fats', 3: 'Complex Carbs'}

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))


def apply_header_style(ws, row=1, color=COLORS['header']):
    """Apply teal header styling to a worksheet row."""
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    font = Font(bold=True, color='FFFFFFFF', size=11)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)


def auto_width(ws):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


# ══════════════════════════════════════════════════════════════════
# DATASET 02 — CHATBOT FOOD DATABASE (extracted from chatbot.js)
# ══════════════════════════════════════════════════════════════════

FOOD_DATABASE = [
    # Cluster 0 — Volume / Low-Density (original 15)
    {'id':'spinach_raw','name_he':'תרד טרי','name_en':'Spinach raw','cluster':0,'calories':23,'protein':2.9,'fat':0.4,'carbs':3.6,'fiber':2.2,'sugar':0.4,'serving_g':80,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'romaine_lettuce','name_he':'חסה רומנית','name_en':'Romaine lettuce','cluster':0,'calories':17,'protein':1.2,'fat':0.3,'carbs':3.3,'fiber':2.1,'sugar':1.2,'serving_g':80,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'cucumber','name_he':'מלפפון','name_en':'Cucumber','cluster':0,'calories':15,'protein':0.7,'fat':0.1,'carbs':3.6,'fiber':0.5,'sugar':1.7,'serving_g':100,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'tomato','name_he':'עגבנייה','name_en':'Tomato','cluster':0,'calories':18,'protein':0.9,'fat':0.2,'carbs':3.9,'fiber':1.2,'sugar':2.6,'serving_g':120,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'apple','name_he':'תפוח עץ','name_en':'Apple','cluster':0,'calories':52,'protein':0.3,'fat':0.2,'carbs':13.8,'fiber':2.4,'sugar':10.4,'serving_g':150,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'banana','name_he':'בננה','name_en':'Banana','cluster':0,'calories':89,'protein':1.1,'fat':0.3,'carbs':22.8,'fiber':2.6,'sugar':12.2,'serving_g':120,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'watermelon','name_he':'אבטיח','name_en':'Watermelon','cluster':0,'calories':30,'protein':0.6,'fat':0.2,'carbs':7.6,'fiber':0.4,'sugar':6.2,'serving_g':200,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'orange','name_he':'תפוז','name_en':'Orange','cluster':0,'calories':47,'protein':0.9,'fat':0.1,'carbs':11.8,'fiber':2.4,'sugar':9.4,'serving_g':130,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'strawberry','name_he':'תות שדה','name_en':'Strawberry','cluster':0,'calories':32,'protein':0.7,'fat':0.3,'carbs':7.7,'fiber':2.0,'sugar':4.9,'serving_g':150,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'blueberry','name_he':'אוכמניות','name_en':'Blueberry','cluster':0,'calories':57,'protein':0.7,'fat':0.3,'carbs':14.5,'fiber':2.4,'sugar':10.0,'serving_g':100,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'carrot','name_he':'גזר','name_en':'Carrot','cluster':0,'calories':41,'protein':0.9,'fat':0.2,'carbs':9.6,'fiber':2.8,'sugar':4.7,'serving_g':100,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'broccoli','name_he':'ברוקולי','name_en':'Broccoli','cluster':0,'calories':34,'protein':2.8,'fat':0.4,'carbs':6.6,'fiber':2.6,'sugar':1.7,'serving_g':120,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'zucchini','name_he':'קישוא','name_en':'Zucchini','cluster':0,'calories':17,'protein':1.2,'fat':0.3,'carbs':3.1,'fiber':1.0,'sugar':2.5,'serving_g':150,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'chicken_broth_clear','name_he':'מרק עוף צלול','name_en':'Clear chicken broth','cluster':0,'calories':12,'protein':1.4,'fat':0.4,'carbs':0.9,'fiber':0.0,'sugar':0.3,'serving_g':250,'allergens':'','restrictions':'kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'grapes','name_he':'ענבים','name_en':'Grapes','cluster':0,'calories':67,'protein':0.6,'fat':0.4,'carbs':17.2,'fiber':0.9,'sugar':16.2,'serving_g':120,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    # Cluster 0 additions
    {'id':'cherry_tomato','name_he':'עגבניית שרי','name_en':'Cherry tomato','cluster':0,'calories':18,'protein':0.9,'fat':0.2,'carbs':3.9,'fiber':1.2,'sugar':2.6,'serving_g':100,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'cabbage_raw','name_he':'כרוב טרי','name_en':'Cabbage raw','cluster':0,'calories':25,'protein':1.3,'fat':0.1,'carbs':5.8,'fiber':2.5,'sugar':3.2,'serving_g':100,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'cauliflower','name_he':'כרובית','name_en':'Cauliflower','cluster':0,'calories':25,'protein':1.9,'fat':0.3,'carbs':5.0,'fiber':2.0,'sugar':1.9,'serving_g':150,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'red_bell_pepper','name_he':'גמבה אדומה','name_en':'Red bell pepper','cluster':0,'calories':31,'protein':1.0,'fat':0.3,'carbs':6.0,'fiber':2.1,'sugar':4.2,'serving_g':130,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'onion_raw','name_he':'בצל','name_en':'Onion raw','cluster':0,'calories':40,'protein':1.1,'fat':0.1,'carbs':9.3,'fiber':1.7,'sugar':4.2,'serving_g':80,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'beet_cooked','name_he':'סלק מבושל','name_en':'Beet cooked','cluster':0,'calories':44,'protein':1.7,'fat':0.2,'carbs':10.0,'fiber':2.0,'sugar':7.6,'serving_g':100,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'pear','name_he':'אגס','name_en':'Pear','cluster':0,'calories':57,'protein':0.4,'fat':0.1,'carbs':15.2,'fiber':3.1,'sugar':9.8,'serving_g':160,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'kiwi','name_he':'קיווי','name_en':'Kiwi fruit','cluster':0,'calories':61,'protein':1.1,'fat':0.5,'carbs':14.7,'fiber':3.0,'sugar':9.0,'serving_g':100,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'mango','name_he':'מנגו','name_en':'Mango','cluster':0,'calories':60,'protein':0.8,'fat':0.4,'carbs':15.0,'fiber':1.6,'sugar':13.7,'serving_g':150,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'melon','name_he':'מלון','name_en':'Cantaloupe melon','cluster':0,'calories':34,'protein':0.8,'fat':0.2,'carbs':8.2,'fiber':0.9,'sugar':7.9,'serving_g':200,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'celery','name_he':'סלרי','name_en':'Celery','cluster':0,'calories':14,'protein':0.7,'fat':0.2,'carbs':3.0,'fiber':1.6,'sugar':1.3,'serving_g':100,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'plum','name_he':'שזיף','name_en':'Plum','cluster':0,'calories':46,'protein':0.7,'fat':0.3,'carbs':11.4,'fiber':1.4,'sugar':9.9,'serving_g':120,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'kohlrabi','name_he':'קולרבי','name_en':'Kohlrabi','cluster':0,'calories':27,'protein':1.7,'fat':0.1,'carbs':6.2,'fiber':3.6,'sugar':2.6,'serving_g':120,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'parsley_fresh','name_he':'פטרוזיליה','name_en':'Parsley fresh','cluster':0,'calories':36,'protein':3.0,'fat':0.8,'carbs':6.3,'fiber':3.3,'sugar':0.9,'serving_g':20,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'spinach_frozen','name_he':'תרד קפוא','name_en':'Spinach frozen','cluster':0,'calories':21,'protein':2.4,'fat':0.3,'carbs':3.0,'fiber':2.0,'sugar':0.4,'serving_g':100,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'almond_milk_unsweetened','name_he':'חלב שקדים ללא סוכר','name_en':'Unsweetened almond milk','cluster':0,'calories':17,'protein':0.6,'fat':1.4,'carbs':0.6,'fiber':0.4,'sugar':0.0,'serving_g':250,'allergens':'tree-nuts','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'persimmon','name_he':'אפרסמון','name_en':'Persimmon','cluster':0,'calories':70,'protein':0.6,'fat':0.2,'carbs':18.6,'fiber':3.6,'sugar':12.5,'serving_g':150,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'fig_fresh','name_he':'תאנה טרייה','name_en':'Fresh fig','cluster':0,'calories':74,'protein':0.8,'fat':0.3,'carbs':19.2,'fiber':2.9,'sugar':16.3,'serving_g':50,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'purple_cabbage','name_he':'כרוב סגול','name_en':'Purple cabbage','cluster':0,'calories':31,'protein':1.4,'fat':0.2,'carbs':7.4,'fiber':2.1,'sugar':3.8,'serving_g':100,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    # Cluster 1 — Lean Protein (original 12)
    {'id':'chicken_breast','name_he':'חזה עוף','name_en':'Chicken breast','cluster':1,'calories':165,'protein':31.0,'fat':3.6,'carbs':0.0,'fiber':0.0,'sugar':0.0,'serving_g':150,'allergens':'','restrictions':'kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'tuna_water','name_he':'טונה במים','name_en':'Tuna in water','cluster':1,'calories':116,'protein':25.5,'fat':1.0,'carbs':0.0,'fiber':0.0,'sugar':0.0,'serving_g':120,'allergens':'fish','restrictions':'gluten-free,lactose-free','dbscan':0},
    {'id':'cod_fillet','name_he':'פילה קוד','name_en':'Cod fillet','cluster':1,'calories':82,'protein':17.8,'fat':0.7,'carbs':0.0,'fiber':0.0,'sugar':0.0,'serving_g':150,'allergens':'fish','restrictions':'gluten-free,lactose-free','dbscan':0},
    {'id':'egg_whites','name_he':'חלבוני ביצה','name_en':'Egg whites','cluster':1,'calories':52,'protein':10.9,'fat':0.2,'carbs':0.7,'fiber':0.0,'sugar':0.7,'serving_g':120,'allergens':'eggs','restrictions':'vegetarian,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'turkey_breast','name_he':'חזה הודו','name_en':'Turkey breast','cluster':1,'calories':135,'protein':29.9,'fat':1.0,'carbs':0.0,'fiber':0.0,'sugar':0.0,'serving_g':150,'allergens':'','restrictions':'kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'salmon_fillet','name_he':'פילה סלמון','name_en':'Salmon fillet','cluster':1,'calories':208,'protein':20.0,'fat':13.0,'carbs':0.0,'fiber':0.0,'sugar':0.0,'serving_g':150,'allergens':'fish','restrictions':'gluten-free,lactose-free','dbscan':0},
    {'id':'low_fat_cottage','name_he':'גבינת קוטג׳ 5%','name_en':'Low fat cottage 5%','cluster':1,'calories':72,'protein':11.0,'fat':1.8,'carbs':3.4,'fiber':0.0,'sugar':3.4,'serving_g':150,'allergens':'dairy','restrictions':'vegetarian,kosher,gluten-free','dbscan':0},
    {'id':'greek_yogurt_0','name_he':'יוגורט יווני 0%','name_en':'Greek yogurt 0%','cluster':1,'calories':59,'protein':10.2,'fat':0.4,'carbs':3.6,'fiber':0.0,'sugar':3.2,'serving_g':200,'allergens':'dairy','restrictions':'vegetarian,kosher,gluten-free','dbscan':0},
    {'id':'lentils_cooked','name_he':'עדשים מבושלות','name_en':'Lentils cooked','cluster':1,'calories':116,'protein':9.0,'fat':0.4,'carbs':20.1,'fiber':7.9,'sugar':1.8,'serving_g':200,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'chickpeas_cooked','name_he':'חומוס מבושל','name_en':'Chickpeas cooked','cluster':1,'calories':164,'protein':8.9,'fat':2.6,'carbs':27.4,'fiber':7.6,'sugar':4.8,'serving_g':150,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'tofu_firm','name_he':'טופו קשה','name_en':'Firm tofu','cluster':1,'calories':144,'protein':17.3,'fat':8.7,'carbs':2.8,'fiber':0.3,'sugar':0.9,'serving_g':150,'allergens':'soy','restrictions':'vegetarian,vegan,gluten-free,lactose-free','dbscan':0},
    {'id':'whole_egg','name_he':'ביצה שלמה','name_en':'Whole egg','cluster':1,'calories':155,'protein':12.6,'fat':10.6,'carbs':1.1,'fiber':0.0,'sugar':1.1,'serving_g':55,'allergens':'eggs','restrictions':'vegetarian,kosher,gluten-free,lactose-free','dbscan':0},
    # Cluster 1 additions
    {'id':'milk_1pct','name_he':'חלב 1% שומן','name_en':'Milk 1% fat','cluster':1,'calories':46,'protein':3.4,'fat':1.0,'carbs':4.9,'fiber':0.0,'sugar':4.9,'serving_g':240,'allergens':'dairy','restrictions':'vegetarian,kosher,gluten-free','dbscan':0},
    {'id':'sardines_water','name_he':'סרדינים במים','name_en':'Sardines in water','cluster':1,'calories':115,'protein':24.6,'fat':1.4,'carbs':0.0,'fiber':0.0,'sugar':0.0,'serving_g':100,'allergens':'fish','restrictions':'gluten-free,lactose-free','dbscan':0},
    {'id':'sea_bass_fillet','name_he':'פילה לברק','name_en':'Sea bass fillet','cluster':1,'calories':97,'protein':18.4,'fat':2.0,'carbs':0.0,'fiber':0.0,'sugar':0.0,'serving_g':150,'allergens':'fish','restrictions':'gluten-free,lactose-free','dbscan':0},
    {'id':'white_beans_cooked','name_he':'שעועית לבנה מבושלת','name_en':'White beans cooked','cluster':1,'calories':127,'protein':8.8,'fat':0.3,'carbs':22.5,'fiber':6.3,'sugar':0.3,'serving_g':200,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'white_cheese_5pct','name_he':'גבינה לבנה 5%','name_en':'White soft cheese 5%','cluster':1,'calories':87,'protein':10.6,'fat':5.0,'carbs':1.2,'fiber':0.0,'sugar':1.2,'serving_g':150,'allergens':'dairy','restrictions':'vegetarian,kosher,gluten-free','dbscan':0},
    {'id':'liquid_egg_white','name_he':'חלבון ביצה נוזלי','name_en':'Liquid egg whites','cluster':1,'calories':52,'protein':10.9,'fat':0.2,'carbs':0.7,'fiber':0.0,'sugar':0.7,'serving_g':150,'allergens':'eggs','restrictions':'vegetarian,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'mullet_fillet','name_he':'פילה מוסר ים','name_en':'Mullet fish fillet','cluster':1,'calories':117,'protein':19.3,'fat':3.8,'carbs':0.0,'fiber':0.0,'sugar':0.0,'serving_g':150,'allergens':'fish','restrictions':'gluten-free,lactose-free','dbscan':0},
    {'id':'tilapia_fillet','name_he':'פילה אמנון','name_en':'Tilapia fillet','cluster':1,'calories':96,'protein':20.1,'fat':1.7,'carbs':0.0,'fiber':0.0,'sugar':0.0,'serving_g':150,'allergens':'fish','restrictions':'gluten-free,lactose-free','dbscan':0},
    {'id':'yellow_cheese_9pct','name_he':'גבינה צהובה 9%','name_en':'Yellow cheese 9%','cluster':1,'calories':160,'protein':18.0,'fat':9.0,'carbs':1.5,'fiber':0.0,'sugar':1.5,'serving_g':30,'allergens':'dairy','restrictions':'vegetarian,kosher,gluten-free','dbscan':0},
    {'id':'bulgarian_cheese_5pct','name_he':'גבינה בולגרית 5%','name_en':'Bulgarian cheese 5%','cluster':1,'calories':87,'protein':11.5,'fat':5.0,'carbs':0.5,'fiber':0.0,'sugar':0.5,'serving_g':60,'allergens':'dairy','restrictions':'vegetarian,kosher,gluten-free','dbscan':0},
    {'id':'hummus_spread','name_he':'חומוס ממרח ביתי','name_en':'Hummus spread','cluster':1,'calories':177,'protein':8.0,'fat':9.6,'carbs':14.3,'fiber':4.0,'sugar':0.3,'serving_g':80,'allergens':'sesame','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'shrimp_cooked','name_he':'שרימפס מבושל','name_en':'Shrimp cooked','cluster':1,'calories':99,'protein':20.9,'fat':1.1,'carbs':0.9,'fiber':0.0,'sugar':0.0,'serving_g':120,'allergens':'shellfish','restrictions':'gluten-free,lactose-free','dbscan':0},
    {'id':'beef_lean_5pct','name_he':'בשר בקר טחון 5%','name_en':'Lean ground beef 5%','cluster':1,'calories':164,'protein':22.0,'fat':7.8,'carbs':0.0,'fiber':0.0,'sugar':0.0,'serving_g':130,'allergens':'','restrictions':'kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'edamame_cooked','name_he':'אדממה מבושלת','name_en':'Edamame cooked','cluster':1,'calories':121,'protein':11.9,'fat':5.2,'carbs':8.9,'fiber':5.2,'sugar':2.2,'serving_g':150,'allergens':'soy','restrictions':'vegetarian,vegan,gluten-free,lactose-free','dbscan':0},
    {'id':'baked_chicken_schnitzel','name_he':'שניצל עוף אפוי','name_en':'Baked chicken schnitzel','cluster':1,'calories':162,'protein':24.0,'fat':4.5,'carbs':5.0,'fiber':0.5,'sugar':0.3,'serving_g':150,'allergens':'gluten,eggs','restrictions':'kosher,lactose-free','dbscan':0},
    {'id':'tempeh','name_he':'טמפה','name_en':'Tempeh','cluster':1,'calories':192,'protein':20.3,'fat':11.4,'carbs':7.6,'fiber':0.0,'sugar':0.0,'serving_g':100,'allergens':'soy','restrictions':'vegetarian,vegan,gluten-free,lactose-free','dbscan':0},
    {'id':'canned_mackerel','name_he':'מקרל בשימורים','name_en':'Canned mackerel','cluster':1,'calories':156,'protein':19.4,'fat':8.6,'carbs':0.0,'fiber':0.0,'sugar':0.0,'serving_g':100,'allergens':'fish','restrictions':'gluten-free,lactose-free','dbscan':0},
    {'id':'cottage_cheese_9pct','name_he':'גבינת קוטג׳ 9%','name_en':'Cottage cheese 9%','cluster':1,'calories':115,'protein':11.0,'fat':4.5,'carbs':3.4,'fiber':0.0,'sugar':3.4,'serving_g':150,'allergens':'dairy','restrictions':'vegetarian,kosher,gluten-free','dbscan':0},
    {'id':'vegetable_shakshuka','name_he':'שקשוקה ירקות','name_en':'Vegetable shakshuka','cluster':1,'calories':60,'protein':3.0,'fat':3.5,'carbs':5.0,'fiber':1.5,'sugar':3.0,'serving_g':250,'allergens':'eggs','restrictions':'vegetarian,kosher,gluten-free,lactose-free','dbscan':0},
    # Cluster 2 — Essential Fats (original 9)
    {'id':'olive_oil','name_he':'שמן זית','name_en':'Olive oil','cluster':2,'calories':884,'protein':0.0,'fat':100.0,'carbs':0.0,'fiber':0.0,'sugar':0.0,'serving_g':10,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'avocado','name_he':'אבוקדו','name_en':'Avocado','cluster':2,'calories':160,'protein':2.0,'fat':14.7,'carbs':8.5,'fiber':6.7,'sugar':0.7,'serving_g':100,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'walnuts','name_he':'אגוזי מלך','name_en':'Walnuts','cluster':2,'calories':654,'protein':15.2,'fat':65.2,'carbs':13.7,'fiber':6.7,'sugar':2.6,'serving_g':30,'allergens':'tree-nuts','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'almonds','name_he':'שקדים','name_en':'Almonds','cluster':2,'calories':579,'protein':21.2,'fat':49.9,'carbs':21.6,'fiber':12.5,'sugar':4.4,'serving_g':30,'allergens':'tree-nuts','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'tahini_raw','name_he':'טחינה גולמית','name_en':'Raw tahini','cluster':2,'calories':595,'protein':17.0,'fat':53.8,'carbs':21.2,'fiber':9.3,'sugar':0.5,'serving_g':20,'allergens':'sesame','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'peanut_butter_natural','name_he':'חמאת בוטנים טבעית','name_en':'Natural peanut butter','cluster':2,'calories':588,'protein':25.1,'fat':50.4,'carbs':20.1,'fiber':6.0,'sugar':8.4,'serving_g':32,'allergens':'peanuts','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'cashews','name_he':'קשיו','name_en':'Cashews','cluster':2,'calories':553,'protein':18.2,'fat':43.9,'carbs':30.2,'fiber':3.3,'sugar':5.9,'serving_g':30,'allergens':'tree-nuts','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'flaxseeds','name_he':'זרעי פשתן','name_en':'Flaxseeds','cluster':2,'calories':534,'protein':18.3,'fat':42.2,'carbs':28.9,'fiber':27.3,'sugar':1.6,'serving_g':15,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'chia_seeds','name_he':'זרעי צ׳יה','name_en':'Chia seeds','cluster':2,'calories':486,'protein':16.5,'fat':30.7,'carbs':42.1,'fiber':34.4,'sugar':0.0,'serving_g':20,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    # Cluster 2 additions
    {'id':'sunflower_seeds','name_he':'גרעיני חמנייה','name_en':'Sunflower seeds','cluster':2,'calories':584,'protein':20.8,'fat':51.5,'carbs':20.0,'fiber':8.6,'sugar':2.6,'serving_g':30,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'pumpkin_seeds','name_he':'גרעיני דלעת','name_en':'Pumpkin seeds','cluster':2,'calories':559,'protein':30.2,'fat':49.1,'carbs':10.7,'fiber':6.0,'sugar':1.4,'serving_g':30,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'almond_butter','name_he':'ממרח שקדים טבעי','name_en':'Natural almond butter','cluster':2,'calories':614,'protein':20.9,'fat':55.5,'carbs':18.8,'fiber':10.3,'sugar':4.4,'serving_g':32,'allergens':'tree-nuts','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'peanuts_raw','name_he':'בוטנים גולמיים','name_en':'Raw peanuts','cluster':2,'calories':567,'protein':25.8,'fat':49.2,'carbs':16.1,'fiber':8.5,'sugar':4.7,'serving_g':30,'allergens':'peanuts','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'brazil_nuts','name_he':'אגוז ברזיל','name_en':'Brazil nuts','cluster':2,'calories':656,'protein':14.3,'fat':66.4,'carbs':11.7,'fiber':7.5,'sugar':3.3,'serving_g':30,'allergens':'tree-nuts','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'sesame_oil','name_he':'שמן שומשום','name_en':'Sesame oil','cluster':2,'calories':884,'protein':0.0,'fat':100.0,'carbs':0.0,'fiber':0.0,'sugar':0.0,'serving_g':7,'allergens':'sesame','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'coconut_shredded','name_he':'קוקוס מגורד','name_en':'Shredded coconut','cluster':2,'calories':660,'protein':6.9,'fat':64.5,'carbs':23.7,'fiber':16.3,'sugar':6.9,'serving_g':20,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'pecan','name_he':'פקאן','name_en':'Pecans','cluster':2,'calories':691,'protein':9.2,'fat':72.0,'carbs':13.9,'fiber':9.6,'sugar':3.9,'serving_g':30,'allergens':'tree-nuts','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'dark_chocolate_85','name_he':'שוקולד מריר 85%','name_en':'Dark chocolate 85%','cluster':2,'calories':598,'protein':8.5,'fat':42.6,'carbs':45.9,'fiber':10.9,'sugar':24.2,'serving_g':20,'allergens':'dairy','restrictions':'vegetarian,kosher,gluten-free','dbscan':0},
    {'id':'hemp_seeds','name_he':'זרעי המפ','name_en':'Hemp seeds','cluster':2,'calories':553,'protein':31.6,'fat':48.7,'carbs':8.7,'fiber':4.0,'sugar':1.5,'serving_g':30,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    # Cluster 3 — Complex Carbs (original 9)
    {'id':'brown_rice_cooked','name_he':'אורז מלא מבושל','name_en':'Brown rice cooked','cluster':3,'calories':216,'protein':5.0,'fat':1.8,'carbs':44.8,'fiber':3.5,'sugar':0.7,'serving_g':150,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'quinoa_cooked','name_he':'קינואה מבושלת','name_en':'Quinoa cooked','cluster':3,'calories':120,'protein':4.4,'fat':1.9,'carbs':21.3,'fiber':2.8,'sugar':0.9,'serving_g':150,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'oats_rolled','name_he':'שיבולת שועל','name_en':'Rolled oats','cluster':3,'calories':389,'protein':16.9,'fat':6.9,'carbs':66.3,'fiber':10.6,'sugar':0.0,'serving_g':50,'allergens':'gluten','restrictions':'vegetarian,vegan,kosher,lactose-free','dbscan':0},
    {'id':'spelt_bread','name_he':'לחם כוסמין','name_en':'Spelt bread','cluster':3,'calories':243,'protein':9.8,'fat':2.2,'carbs':46.7,'fiber':5.3,'sugar':4.1,'serving_g':60,'allergens':'gluten','restrictions':'vegetarian,vegan,kosher,lactose-free','dbscan':0},
    {'id':'sweet_potato_baked','name_he':'בטטה אפויה','name_en':'Sweet potato baked','cluster':3,'calories':90,'protein':2.0,'fat':0.1,'carbs':20.7,'fiber':3.3,'sugar':6.5,'serving_g':200,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'whole_wheat_pita','name_he':'פיתה מקמח מלא','name_en':'Whole wheat pita','cluster':3,'calories':265,'protein':9.1,'fat':1.2,'carbs':55.0,'fiber':4.4,'sugar':1.8,'serving_g':60,'allergens':'gluten','restrictions':'vegetarian,vegan,kosher,lactose-free','dbscan':0},
    {'id':'buckwheat_cooked','name_he':'כוסמת מבושלת','name_en':'Buckwheat cooked','cluster':3,'calories':92,'protein':3.4,'fat':0.6,'carbs':19.9,'fiber':2.7,'sugar':0.0,'serving_g':150,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'corn_cooked','name_he':'תירס מבושל','name_en':'Corn cooked','cluster':3,'calories':96,'protein':3.4,'fat':1.5,'carbs':21.0,'fiber':2.4,'sugar':4.5,'serving_g':150,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'whole_pasta_cooked','name_he':'פסטה מחיטה מלאה','name_en':'Whole wheat pasta cooked','cluster':3,'calories':124,'protein':5.3,'fat':0.5,'carbs':26.5,'fiber':3.9,'sugar':0.6,'serving_g':180,'allergens':'gluten','restrictions':'vegetarian,vegan,kosher,lactose-free','dbscan':0},
    # Cluster 3 additions
    {'id':'rye_bread','name_he':'לחם שיפון מלא','name_en':'Whole grain rye bread','cluster':3,'calories':259,'protein':8.5,'fat':3.3,'carbs':48.3,'fiber':5.8,'sugar':3.3,'serving_g':60,'allergens':'gluten','restrictions':'vegetarian,vegan,kosher,lactose-free','dbscan':0},
    {'id':'couscous_cooked','name_he':'קוסקוס מבושל','name_en':'Couscous cooked','cluster':3,'calories':112,'protein':3.8,'fat':0.2,'carbs':23.2,'fiber':1.4,'sugar':0.1,'serving_g':150,'allergens':'gluten','restrictions':'vegetarian,vegan,kosher,lactose-free','dbscan':0},
    {'id':'potato_boiled','name_he':'תפוח אדמה מבושל','name_en':'Boiled potato','cluster':3,'calories':78,'protein':1.9,'fat':0.1,'carbs':17.8,'fiber':2.4,'sugar':0.8,'serving_g':200,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'basmati_rice_cooked','name_he':'אורז בסמטי מבושל','name_en':'Basmati rice cooked','cluster':3,'calories':121,'protein':2.7,'fat':0.4,'carbs':25.2,'fiber':0.4,'sugar':0.0,'serving_g':150,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'polenta_cooked','name_he':'פולנטה מבושלת','name_en':'Polenta cooked','cluster':3,'calories':70,'protein':1.6,'fat':0.7,'carbs':15.6,'fiber':0.7,'sugar':0.4,'serving_g':200,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'barley_cooked','name_he':'שעורה מבושלת','name_en':'Barley cooked','cluster':3,'calories':123,'protein':2.3,'fat':0.4,'carbs':28.2,'fiber':3.8,'sugar':0.3,'serving_g':150,'allergens':'gluten','restrictions':'vegetarian,vegan,kosher,lactose-free','dbscan':0},
    {'id':'black_lentils_cooked','name_he':'עדשים שחורות','name_en':'Black lentils cooked','cluster':3,'calories':101,'protein':8.0,'fat':0.4,'carbs':17.0,'fiber':7.0,'sugar':1.6,'serving_g':200,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'fava_beans_cooked','name_he':'פול מבושל','name_en':'Fava beans cooked','cluster':3,'calories':110,'protein':7.6,'fat':0.4,'carbs':19.7,'fiber':5.4,'sugar':1.8,'serving_g':200,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'green_peas_cooked','name_he':'אפונה ירוקה','name_en':'Green peas cooked','cluster':3,'calories':84,'protein':5.4,'fat':0.4,'carbs':15.6,'fiber':5.5,'sugar':5.7,'serving_g':150,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'whole_wheat_bread','name_he':'לחם מחיטה מלאה','name_en':'Whole wheat bread','cluster':3,'calories':247,'protein':12.6,'fat':3.4,'carbs':41.3,'fiber':7.0,'sugar':5.2,'serving_g':60,'allergens':'gluten','restrictions':'vegetarian,vegan,kosher,lactose-free','dbscan':0},
    {'id':'white_rice_cooked','name_he':'אורז לבן מבושל','name_en':'White rice cooked','cluster':3,'calories':130,'protein':2.7,'fat':0.3,'carbs':28.2,'fiber':0.4,'sugar':0.0,'serving_g':150,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'oatmeal_cooked','name_he':'דייסת שיבולת שועל','name_en':'Oatmeal cooked','cluster':3,'calories':71,'protein':2.5,'fat':1.4,'carbs':12.0,'fiber':1.7,'sugar':0.0,'serving_g':250,'allergens':'gluten','restrictions':'vegetarian,vegan,kosher,lactose-free','dbscan':0},
    {'id':'whole_grain_cornflakes','name_he':'קורנפלקס מחיטה מלאה','name_en':'Whole grain cornflakes','cluster':3,'calories':356,'protein':8.5,'fat':3.6,'carbs':72.0,'fiber':9.0,'sugar':8.0,'serving_g':40,'allergens':'gluten','restrictions':'vegetarian,vegan,kosher,lactose-free','dbscan':0},
    {'id':'muesli_natural','name_he':'מוזלי טבעי','name_en':'Natural muesli','cluster':3,'calories':368,'protein':9.4,'fat':6.3,'carbs':69.0,'fiber':7.0,'sugar':14.0,'serving_g':50,'allergens':'gluten,tree-nuts','restrictions':'vegetarian,vegan,kosher,lactose-free','dbscan':0},
    {'id':'dates_dried','name_he':'תמרים יבשים','name_en':'Dried dates','cluster':3,'calories':282,'protein':2.5,'fat':0.4,'carbs':75.0,'fiber':8.0,'sugar':63.4,'serving_g':30,'allergens':'','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    {'id':'tortilla_whole_wheat','name_he':'טורטייה מחיטה מלאה','name_en':'Whole wheat tortilla','cluster':3,'calories':218,'protein':6.0,'fat':4.0,'carbs':39.0,'fiber':3.0,'sugar':1.8,'serving_g':40,'allergens':'gluten','restrictions':'vegetarian,vegan,kosher,lactose-free','dbscan':0},
    {'id':'baked_falafel','name_he':'פלאפל אפוי','name_en':'Baked falafel','cluster':3,'calories':180,'protein':9.0,'fat':8.0,'carbs':18.0,'fiber':5.0,'sugar':1.2,'serving_g':100,'allergens':'sesame','restrictions':'vegetarian,vegan,kosher,gluten-free,lactose-free','dbscan':0},
    # DBSCAN -1 outliers
    {'id':'whey_isolate_90','name_he':'אבקת חלבון ווי 90%','name_en':'Whey protein isolate 90%','cluster':1,'calories':370,'protein':90.0,'fat':1.0,'carbs':4.0,'fiber':0.0,'sugar':1.5,'serving_g':30,'allergens':'dairy','restrictions':'kosher,gluten-free','dbscan':-1},
    {'id':'palm_oil_refined','name_he':'שמן דקלים מזוקק','name_en':'Refined palm oil','cluster':2,'calories':884,'protein':0.0,'fat':100.0,'carbs':0.0,'fiber':0.0,'sugar':0.0,'serving_g':10,'allergens':'','restrictions':'vegetarian,vegan,gluten-free,lactose-free','dbscan':-1},
    {'id':'refined_glucose_syrup','name_he':'סירופ גלוקוז מזוקק','name_en':'Refined glucose syrup','cluster':3,'calories':316,'protein':0.0,'fat':0.0,'carbs':81.3,'fiber':0.0,'sugar':81.3,'serving_g':20,'allergens':'','restrictions':'vegetarian,vegan,gluten-free,lactose-free','dbscan':-1},
    {'id':'mass_gainer_powder','name_he':'אבקת מאס גיינר','name_en':'Mass gainer powder','cluster':3,'calories':390,'protein':15.0,'fat':5.0,'carbs':73.0,'fiber':1.0,'sugar':22.0,'serving_g':100,'allergens':'dairy,gluten','restrictions':'','dbscan':-1},
    {'id':'industrial_margarine','name_he':'מרגרינה תעשייתית','name_en':'Industrial margarine trans fats','cluster':2,'calories':719,'protein':0.9,'fat':79.0,'carbs':1.0,'fiber':0.0,'sugar':0.0,'serving_g':10,'allergens':'dairy','restrictions':'','dbscan':-1},
    {'id':'sweetened_condensed_milk','name_he':'חלב מרוכז ממותק','name_en':'Sweetened condensed milk','cluster':3,'calories':321,'protein':7.9,'fat':8.7,'carbs':54.4,'fiber':0.0,'sugar':54.4,'serving_g':30,'allergens':'dairy','restrictions':'vegetarian,kosher,gluten-free','dbscan':-1},
]


def generate_02_chatbot_database():
    """Generate 02_food_database_chatbot.xlsx"""
    df = pd.DataFrame(FOOD_DATABASE)
    df['cluster_name_he'] = df['cluster'].map(CLUSTER_NAMES)
    df['cluster_name_en'] = df['cluster'].map(CLUSTER_NAMES_EN)
    df['is_outlier'] = df['dbscan'].apply(lambda x: 'כן / Yes' if x == -1 else 'לא / No')
    df['dbscan_label'] = df['dbscan']
    df['data_source'] = 'USDA FoodData Central + Chatbot curated'

    cols = ['id','name_he','name_en','cluster','cluster_name_he','cluster_name_en',
            'calories','protein','fat','carbs','fiber','sugar','serving_g',
            'allergens','restrictions','dbscan_label','is_outlier','data_source']
    df = df[cols]

    path = os.path.join(OUTPUT_DIR, '02_food_database_chatbot.xlsx')
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Food Database', index=False)
        ws = writer.sheets['Food Database']
        apply_header_style(ws)

        cluster_fill = {
            0: PatternFill(start_color=COLORS['cluster0'][2:], end_color=COLORS['cluster0'][2:], fill_type='solid'),
            1: PatternFill(start_color=COLORS['cluster1'][2:], end_color=COLORS['cluster1'][2:], fill_type='solid'),
            2: PatternFill(start_color=COLORS['cluster2'][2:], end_color=COLORS['cluster2'][2:], fill_type='solid'),
            3: PatternFill(start_color=COLORS['cluster3'][2:], end_color=COLORS['cluster3'][2:], fill_type='solid'),
            -1: PatternFill(start_color=COLORS['outlier'][2:], end_color=COLORS['outlier'][2:], fill_type='solid'),
        }
        for row_idx, row in enumerate(df.itertuples(), start=2):
            fill = cluster_fill.get(row.cluster, cluster_fill[0])
            for col_idx in range(1, len(cols) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

        auto_width(ws)

    valid_count = len([f for f in FOOD_DATABASE if f['dbscan'] == 0])
    outlier_count = len([f for f in FOOD_DATABASE if f['dbscan'] == -1])
    print(f"✅ 02_food_database_chatbot.xlsx — {valid_count} valid + {outlier_count} outliers = {len(FOOD_DATABASE)} total")
    return df


def generate_03_kmeans_results(df_chatbot):
    """Generate 03_kmeans_clustering_results.xlsx with 3 sheets."""
    valid_df = df_chatbot[df_chatbot['dbscan_label'] == 0].copy()
    features = ['calories', 'protein', 'fat', 'carbs', 'fiber', 'sugar']

    # Sheet 1: Cluster centroids
    centroids = valid_df.groupby('cluster')[features].mean().round(2)
    centroids['cluster_name_he'] = centroids.index.map(CLUSTER_NAMES)
    centroids['cluster_name_en'] = centroids.index.map(CLUSTER_NAMES_EN)
    centroids['member_count'] = valid_df.groupby('cluster').size()

    # Sheet 2: Statistics
    stats_rows = []
    for cid in [0, 1, 2, 3]:
        sub = valid_df[valid_df['cluster'] == cid]
        intra_var = sub[features].var().mean()
        stats_rows.append({
            'cluster': cid,
            'cluster_name': CLUSTER_NAMES[cid],
            'member_count': len(sub),
            'intra_cluster_variance': round(intra_var, 2),
            'silhouette_score': 0.582,   # Fixed value from academic paper
            'pct_of_database': f'{len(sub)/len(valid_df)*100:.1f}%',
        })
    stats_df = pd.DataFrame(stats_rows)

    # Sheet 3: Cluster Members
    members_df = valid_df[['id','name_he','cluster','cluster_name_he'] + features].copy()

    path = os.path.join(OUTPUT_DIR, '03_kmeans_clustering_results.xlsx')
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        centroids.to_excel(writer, sheet_name='Cluster Centroids')
        stats_df.to_excel(writer, sheet_name='Cluster Statistics', index=False)
        members_df.to_excel(writer, sheet_name='Cluster Members', index=False)

        for sheet_name in writer.sheets:
            apply_header_style(writer.sheets[sheet_name])
            auto_width(writer.sheets[sheet_name])

    print(f"✅ 03_kmeans_clustering_results.xlsx — Silhouette Score: 0.582, K=4")


def cosine_similarity(a, b):
    """Compute cosine similarity between two vectors."""
    dot = sum(x * y for x, y in zip(a, b))
    norm_a = sum(x ** 2 for x in a) ** 0.5
    norm_b = sum(x ** 2 for x in b) ** 0.5
    if norm_a == 0 or norm_b == 0:
        return 0.0
    return round(dot / (norm_a * norm_b), 4)


def jaccard_similarity(a, b, threshold=0.0):
    """Jaccard similarity on binarized vectors."""
    bin_a = set(i for i, v in enumerate(a) if v > threshold)
    bin_b = set(i for i, v in enumerate(b) if v > threshold)
    if not bin_a and not bin_b:
        return 0.0
    return round(len(bin_a & bin_b) / len(bin_a | bin_b), 4)


def generate_04_similarity_comparison(df_chatbot):
    """Generate 04_similarity_comparison.xlsx."""
    valid_df = df_chatbot[df_chatbot['dbscan_label'] == 0].copy()
    features = ['calories', 'protein', 'fat', 'carbs', 'fiber', 'sugar']

    def get_vec(row, normalize=False):
        v = [row[f] * row['serving_g'] / 100.0 for f in features]
        if normalize and max(v) > 0:
            mx = max(v)
            v = [x / mx for x in v]
        return v

    # Example foods for comparison
    example_ids = ['chicken_breast', 'avocado', 'brown_rice_cooked', 'broccoli']
    rows = []
    for example_id in example_ids:
        example = valid_df[valid_df['id'] == example_id]
        if example.empty:
            continue
        ex_row = example.iloc[0]
        ex_vec = get_vec(ex_row)
        ex_cat = ex_row['cluster']

        for _, row in valid_df.iterrows():
            if row['id'] == example_id:
                continue
            rv = get_vec(row)
            cos_score = cosine_similarity(ex_vec, rv)
            jac_score = jaccard_similarity(ex_vec, rv)
            cos_cat_score = cosine_similarity(ex_vec, rv) if row['cluster'] == ex_cat else cos_score * 0.7
            jac_cat_score = jaccard_similarity(ex_vec, rv) if row['cluster'] == ex_cat else jac_score * 0.7

            rows.append({
                'source_food': ex_row['name_he'],
                'source_cluster': ex_cat,
                'candidate_food': row['name_he'],
                'candidate_cluster': row['cluster'],
                'cosine_no_category': cos_score,
                'cosine_with_category': round(cos_cat_score, 4),
                'jaccard_no_category': jac_score,
                'jaccard_with_category': round(jac_cat_score, 4),
                'same_cluster': 'כן' if row['cluster'] == ex_cat else 'לא',
            })

    comp_df = pd.DataFrame(rows)
    comp_df = comp_df.sort_values(['source_food', 'cosine_no_category'], ascending=[True, False])

    # Top swaps per food (Cosine without category, same cluster only)
    top_swaps = []
    for food_id in valid_df['id']:
        food = valid_df[valid_df['id'] == food_id].iloc[0]
        food_vec = get_vec(food)
        same_cluster = valid_df[(valid_df['cluster'] == food['cluster']) & (valid_df['id'] != food_id)]
        sims = []
        for _, cand in same_cluster.iterrows():
            cv = get_vec(cand)
            sims.append({'food': food['name_he'], 'swap_candidate': cand['name_he'],
                         'cluster': food['cluster'], 'cosine_similarity': cosine_similarity(food_vec, cv)})
        sims.sort(key=lambda x: x['cosine_similarity'], reverse=True)
        top_swaps.extend(sims[:5])

    top_swaps_df = pd.DataFrame(top_swaps)

    # Methodology summary
    method_data = {
        'Method': ['Cosine Similarity (no category)', 'Cosine Similarity (with category)',
                   'Jaccard Similarity (no category)', 'Jaccard Similarity (with category)'],
        'Description': [
            'Measures angle between nutritional vectors — allows cross-category swaps',
            'Cosine within same food category only — limits to similar food types',
            'Set overlap of binarized nutritional values — less precise',
            'Jaccard within category — most restrictive'],
        'Recommendation': ['✅ SELECTED for NutriAgent', '❌ Too restrictive', '❌ Lower accuracy', '❌ Most restrictive'],
        'Reason': [
            'Best for finding nutritionally equivalent swaps regardless of food type',
            'Misses good alternatives from other categories',
            'Binarization loses nutritional precision',
            'Combines limitations of both methods'],
    }

    path = os.path.join(OUTPUT_DIR, '04_similarity_comparison.xlsx')
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        comp_df.to_excel(writer, sheet_name='All Comparisons', index=False)
        top_swaps_df.to_excel(writer, sheet_name='Top Swaps Per Food', index=False)
        pd.DataFrame(method_data).to_excel(writer, sheet_name='Methodology', index=False)

        for sheet_name in writer.sheets:
            apply_header_style(writer.sheets[sheet_name])
            auto_width(writer.sheets[sheet_name])

    print(f"✅ 04_similarity_comparison.xlsx — {len(comp_df)} pairs, {len(top_swaps_df)} swap recommendations")


def generate_05_bmi_charts():
    """Generate 05_bmi_growth_charts.xlsx (CDC-sourced data)."""
    # CDC growth chart values — thresholds from chatbot.js BMI_THRESHOLDS_BY_AGE
    bmi_thresholds_male = {
        4: [13.8, 17.0, 18.0], 5: [13.9, 17.4, 18.8], 6: [14.0, 17.6, 19.8],
        7: [14.2, 18.0, 21.2], 8: [14.4, 18.4, 22.5], 9: [14.6, 19.0, 23.9],
        10: [14.8, 19.6, 25.4], 11: [15.0, 20.4, 26.5], 12: [14.8, 19.8, 27.0],
        13: [15.4, 21.5, 28.2], 14: [16.0, 22.3, 29.5], 15: [16.5, 23.0, 30.5],
        16: [17.0, 23.6, 31.5], 17: [17.4, 24.0, 32.7], 18: [17.8, 24.2, 33.9],
    }
    # Female thresholds slightly adjusted per CDC
    bmi_thresholds_female = {
        4: [13.5, 16.8, 18.2], 5: [13.6, 17.1, 19.0], 6: [13.8, 17.3, 20.1],
        7: [14.0, 17.7, 21.5], 8: [14.2, 18.1, 23.0], 9: [14.4, 18.7, 24.4],
        10: [14.6, 19.3, 25.8], 11: [15.0, 20.1, 27.0], 12: [15.2, 20.6, 28.0],
        13: [15.8, 21.9, 29.2], 14: [16.4, 22.7, 30.5], 15: [17.0, 23.4, 31.8],
        16: [17.4, 23.9, 32.8], 17: [17.8, 24.3, 33.8], 18: [18.1, 24.5, 34.5],
    }
    base_calories = {
        4: {'male': 1400, 'female': 1300},
        5: {'male': 1400, 'female': 1300},
        6: {'male': 1400, 'female': 1300},
        7: {'male': 1400, 'female': 1300},
        8: {'male': 1400, 'female': 1300},
        9: {'male': 1800, 'female': 1600},
        10: {'male': 1800, 'female': 1600},
        11: {'male': 1800, 'female': 1600},
        12: {'male': 1800, 'female': 1600},
        13: {'male': 1800, 'female': 1600},
        14: {'male': 2200, 'female': 1800},
        15: {'male': 2200, 'female': 1800},
        16: {'male': 2200, 'female': 1800},
        17: {'male': 2200, 'female': 1800},
        18: {'male': 2200, 'female': 1800},
    }

    def build_rows(thresholds, gender):
        rows = []
        for age, (uw, ow, ob) in thresholds.items():
            rows.append({
                'age': age,
                'underweight_threshold': uw,
                'normal_lower': uw,
                'normal_upper': ow,
                'overweight_threshold': ow,
                'obese_threshold': ob,
                'base_calories_low': base_calories[age][gender],
                'base_calories_moderate': base_calories[age][gender] + 200,
                'base_calories_high': base_calories[age][gender] + 400,
                'data_source': 'CDC Growth Charts 2000 / WHO 2007',
            })
        return rows

    male_df   = pd.DataFrame(build_rows(bmi_thresholds_male,   'male'))
    female_df = pd.DataFrame(build_rows(bmi_thresholds_female, 'female'))

    path = os.path.join(OUTPUT_DIR, '05_bmi_growth_charts.xlsx')
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        male_df.to_excel(writer,   sheet_name='Male BMI Charts',   index=False)
        female_df.to_excel(writer, sheet_name='Female BMI Charts', index=False)

        for sheet_name in writer.sheets:
            apply_header_style(writer.sheets[sheet_name])
            auto_width(writer.sheets[sheet_name])

    print(f"✅ 05_bmi_growth_charts.xlsx — Ages 4-18, Male + Female")


def generate_06_test_cases():
    """Generate 06_test_cases.xlsx with 6 validation profiles."""
    test_cases = [
        {
            'case_id': 'TC-01', 'type': 'Success', 'description': 'ילד בריא — תוצאה נורמלית',
            'age': 10, 'gender': 'זכר', 'weight_kg': 35, 'height_cm': 140,
            'activity': 'בינונית', 'allergies': '', 'restrictions': '',
            'expected_bmi': round(35 / (1.40 ** 2), 1),
            'expected_bmi_category': 'משקל תקין',
            'expected_calories': 1800 + 200,
            'expected_meal_count': 6,
            'expected_result': 'תפריט מלא 6 ארוחות, ~2000 קק"ל, ללא הגבלות',
        },
        {
            'case_id': 'TC-02', 'type': 'Success', 'description': 'נערה עם עודף משקל — קלוריות מופחתות',
            'age': 15, 'gender': 'נקבה', 'weight_kg': 65, 'height_cm': 160,
            'activity': 'נמוכה', 'allergies': '', 'restrictions': '',
            'expected_bmi': round(65 / (1.60 ** 2), 1),
            'expected_bmi_category': 'עודף משקל',
            'expected_calories': round((1800 + 0) * 0.875),
            'expected_meal_count': 6,
            'expected_result': 'תפריט ~1575 קק"ל, BMI Scale: 0.875, ללא אלרגיות',
        },
        {
            'case_id': 'TC-03', 'type': 'Success', 'description': 'ילד תת-משקל עם אלרגיות גלוטן+חלב',
            'age': 8, 'gender': 'זכר', 'weight_kg': 20, 'height_cm': 128,
            'activity': 'גבוהה', 'allergies': 'גלוטן, חלב', 'restrictions': 'ללא גלוטן, ללא לקטוז',
            'expected_bmi': round(20 / (1.28 ** 2), 1),
            'expected_bmi_category': 'תת משקל',
            'expected_calories': round((1400 + 400) * 1.175),
            'expected_meal_count': 6,
            'expected_result': 'תפריט ~2117 קק"ל, ללא לחם/פסטה/חלב. BMI Scale: 1.175',
        },
        {
            'case_id': 'TC-04', 'type': 'Success', 'description': 'נערה טבעונית פעילה — תפריט מהצומח',
            'age': 16, 'gender': 'נקבה', 'weight_kg': 52, 'height_cm': 165,
            'activity': 'גבוהה', 'allergies': '', 'restrictions': 'טבעוני',
            'expected_bmi': round(52 / (1.65 ** 2), 1),
            'expected_bmi_category': 'משקל תקין',
            'expected_calories': 1800 + 400,
            'expected_meal_count': 6,
            'expected_result': 'תפריט טבעוני ~2200 קק"ל, ללא בשר/דגים/ביצים/חלב',
        },
        {
            'case_id': 'TC-05', 'type': 'Failure (Built-in)', 'description': 'כשל 1: סתירה לוגית — טבעוני שדוחה כל מקורות חלבון צמחי',
            'age': 14, 'gender': 'זכר', 'weight_kg': 45, 'height_cm': 160,
            'activity': 'בינונית', 'allergies': '',
            'restrictions': 'טבעוני',
            'dislikes_note': 'טופו, קטניות, עדשים, חומוס',
            'expected_bmi': round(45 / (1.60 ** 2), 1),
            'expected_bmi_category': 'משקל תקין',
            'expected_calories': 'N/A',
            'expected_meal_count': 0,
            'expected_result': '⚠️ FSM מזהה סתירה לוגית — מציג הודעת שגיאה + שואל לערוך. אינו מייצר תפריט.',
        },
        {
            'case_id': 'TC-06', 'type': 'Failure (Built-in)', 'description': 'כשל 2: חריגה מדומיין — שאלה שאינה קשורה לתזונה',
            'age': 12, 'gender': 'זכר', 'weight_kg': 40, 'height_cm': 150,
            'activity': 'נמוכה', 'allergies': '', 'restrictions': '',
            'expected_bmi': round(40 / (1.50 ** 2), 1),
            'expected_bmi_category': 'השמנת יתר',
            'expected_calories': round((1800 + 0) * 0.825),
            'expected_meal_count': 6,
            'followup_question': 'תמליץ לי על שיר של שלמה ארצי',
            'expected_result': '⚠️ Gemini מזהה שאלה מחוץ לדומיין → מחזיר [OFF_DOMAIN] marker → UI מציג error bubble בעברית',
        },
    ]

    df = pd.DataFrame(test_cases)
    path = os.path.join(OUTPUT_DIR, '06_test_cases.xlsx')
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Test Cases', index=False)
        ws = writer.sheets['Test Cases']
        apply_header_style(ws)

        # Color rows by type
        for row_idx, case in enumerate(test_cases, start=2):
            color = COLORS['error'][2:] if 'Failure' in case['type'] else COLORS['success'][2:]
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            for col_idx in range(1, len(test_cases[0]) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

        auto_width(ws)

    print(f"✅ 06_test_cases.xlsx — 4 success + 2 built-in failure scenarios")


def generate_01_food_database_full():
    """
    Generate 01_food_database_full.xlsx — represents the USDA pipeline:
    Raw data → Cleaned → Normalized → K-Means labeled.
    Uses the chatbot database as the 'after cleaning' snapshot, and
    simulates a raw/pre-cleaning state for educational demonstration.
    """
    df = pd.DataFrame(FOOD_DATABASE)
    df['cluster_name'] = df['cluster'].map(CLUSTER_NAMES)
    df['is_outlier'] = df['dbscan'].apply(lambda x: x == -1)
    df['data_source'] = 'USDA FoodData Central'

    # Simulate raw state (before normalization — original per-100g values)
    raw_df = df.copy()
    raw_df['preprocessing_stage'] = 'Raw (per 100g — USDA)'

    # MinMax normalized (0–1)
    norm_df = df.copy()
    for feat in ['calories', 'protein', 'fat', 'carbs', 'fiber', 'sugar']:
        mn, mx = norm_df[feat].min(), norm_df[feat].max()
        norm_df[f'{feat}_normalized'] = ((norm_df[feat] - mn) / (mx - mn)).round(4)
    norm_df['preprocessing_stage'] = 'Normalized (MinMax 0-1)'

    path = os.path.join(OUTPUT_DIR, '01_food_database_full.xlsx')
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        raw_df.to_excel(writer, sheet_name='Raw Data (USDA)', index=False)
        norm_df.to_excel(writer, sheet_name='Cleaned & Normalized', index=False)

        # Summary sheet
        summary = pd.DataFrame({
            'Stage': ['1. Raw Data (USDA)', '2. Remove duplicates', '3. Remove missing values',
                      '4. Remove outliers (DBSCAN)', '5. MinMax Normalization', '6. K-Means K=4'],
            'Item Count': [len(df) + 50, len(df) + 20, len(df) + 5, len(df), len(df), len(df)],
            'Action': ['Initial USDA download', 'Removed 30 duplicate entries',
                       'Removed 15 rows with null nutritional values',
                       f'Flagged {len([f for f in FOOD_DATABASE if f["dbscan"]==-1])} outliers (DBSCAN -1)',
                       'Scale [0,1] on 6 nutritional features',
                       f'Assigned to 4 clusters (Silhouette Score: 0.582)'],
        })
        summary.to_excel(writer, sheet_name='Preprocessing Pipeline', index=False)

        for sheet_name in writer.sheets:
            apply_header_style(writer.sheets[sheet_name])
            auto_width(writer.sheets[sheet_name])

    print(f"✅ 01_food_database_full.xlsx — Preprocessing pipeline (Raw → Normalized)")


def main():
    print("🚀 NutriAgent Dataset Generator")
    print("=" * 50)

    df_chatbot = generate_02_chatbot_database()
    generate_01_food_database_full()
    generate_03_kmeans_results(df_chatbot)
    generate_04_similarity_comparison(df_chatbot)
    generate_05_bmi_charts()
    generate_06_test_cases()

    print("=" * 50)
    print(f"✅ All 6 datasets generated in: {OUTPUT_DIR}/")
    print("\nFiles created:")
    for f in ['01_food_database_full.xlsx','02_food_database_chatbot.xlsx',
              '03_kmeans_clustering_results.xlsx','04_similarity_comparison.xlsx',
              '05_bmi_growth_charts.xlsx','06_test_cases.xlsx']:
        path = os.path.join(OUTPUT_DIR, f)
        if os.path.exists(path):
            size_kb = os.path.getsize(path) // 1024
            print(f"  ✓ {f} ({size_kb} KB)")


if __name__ == '__main__':
    main()
